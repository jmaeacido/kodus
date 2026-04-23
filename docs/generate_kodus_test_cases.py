from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = Path(r"C:\Users\jmaeacido\Downloads\Test Case Template.xlsx")
OUTPUT = ROOT / "docs" / "KODUS_Test_Cases.xlsx"


def lines(*items):
    return "\n".join([item for item in items if item])


CASES = []


def add(module, description, preconditions, steps, expected, severity="Major", priority="Medium Priority = 3-5 days"):
    procedure = lines(
        f"Description: {description}",
        f"Preconditions: {preconditions}" if preconditions else "",
        "Steps:",
        *[f"{i + 1}. {step}" for i, step in enumerate(steps)],
    )
    expected_result = lines(*[f"{i + 1}. {item}" for i, item in enumerate(expected)])
    CASES.append(
        {
            "module": module,
            "procedure": procedure,
            "expected": expected_result,
            "severity": severity,
            "priority": priority,
        }
    )


add("Authentication - Select Year", "Verify the public year selector is required before protected workspace access.", "Browser can reach the public KODUS landing page.", ["Open the public landing page.", "Review the available fiscal year options.", "Select a fiscal year.", "Continue to the login flow."], ["The public year selection page loads without authentication.", "Available fiscal years are displayed.", "The selected year is stored in session and the user is forwarded to sign-in."], "Major", "High Priority= 1-2 days")
add("Authentication - Login", "Verify a valid account can sign in to KODUS.", "A non-deleted account exists and a fiscal year has been selected.", ["Open the login page.", "Enter a valid username or email and password.", "Submit the form.", "Complete 2FA when the account requires it."], ["Valid credentials are accepted.", "The user proceeds to dashboard access or 2FA verification.", "Authenticated session values are created."], "Critical", "High Priority= 1-2 days")
add("Authentication - Login Validation", "Verify login rejects an invalid password.", "A valid account exists.", ["Open the login page.", "Enter a valid username with an invalid password.", "Submit the form."], ["Authentication fails.", "The user remains outside the protected workspace.", "An error is shown for invalid credentials."], "Critical", "High Priority= 1-2 days")
add("Authentication - Registration", "Verify public registration creates a standard KODUS account.", "The target username and email are unused.", ["Open the registration page.", "Enter all required profile fields.", "Enter a strong password and matching confirmation.", "Submit the registration form."], ["The account is created successfully.", "The user record is stored with default picture and standard role behavior.", "A success prompt redirects the user to sign in."])
add("Authentication - Registration Validation", "Verify registration blocks duplicate username or email values.", "An account already exists with the target username or email.", ["Open the registration page.", "Enter a duplicate username or email.", "Complete the remaining required fields.", "Submit the form."], ["Registration is rejected.", "The user is told the username or email is already registered.", "No duplicate account is created."])
add("Authentication - Password Reset", "Verify a user can reset the password with a valid reset token.", "A valid reset token has been issued for an existing account.", ["Open the reset password page with the token.", "Enter a strong replacement password.", "Submit the reset form.", "Sign in using the new password."], ["The token is accepted.", "The password is updated successfully.", "The new password allows login."], "Critical", "High Priority= 1-2 days")
add("Authentication - Logout", "Verify logout ends the authenticated session.", "An account is signed in.", ["Click the logout action.", "Allow the redirect to complete.", "Try to reopen a protected page directly."], ["Session cleanup runs.", "The user returns to the public landing page.", "Protected pages require sign-in again."], "Critical", "High Priority= 1-2 days")
add("Authentication - Legacy Endpoint", "Verify the legacy ajax_login endpoint is disabled.", "None.", ["Send a request to ajax_login.php.", "Inspect the status and response body."], ["The endpoint returns the disabled response.", "The response instructs clients to use the standard login form.", "No session is created through the legacy endpoint."], "Moderate", "Low Priority = 7-10 days")
add("Dashboard - Role Workspace", "Verify dashboard actions reflect the current role.", "Accounts exist for Administrator, AA, Implementation Editor, and User.", ["Sign in as each role.", "Open the Dashboard.", "Review hero actions, quick actions, and workspace summaries."], ["The dashboard loads for each authenticated role.", "Visible shortcuts differ by implemented role checks.", "Restricted modules are not shown to unauthorized roles."])
add("Settings - Profile Update", "Verify an authenticated user can save valid profile changes.", "The user is signed in.", ["Open Settings.", "Update required profile fields with valid values.", "Optionally upload a valid JPG or PNG profile picture.", "Submit the form."], ["The profile changes are saved.", "Updated values persist in the users table.", "A success flash message is shown."])
add("Settings - Validation", "Verify Settings rejects incomplete or invalid profile data.", "The user is signed in.", ["Open Settings.", "Remove a required field or enter an invalid email value.", "Submit the form."], ["The update is rejected.", "The user remains in the settings workflow.", "An error flash explains the validation problem."])
add("Settings - Password Policy Validation", "Verify Settings blocks a weak replacement password.", "The user is signed in.", ["Open Settings.", "Enter a password that does not meet the policy.", "Submit the form."], ["The password is not updated.", "A weak password warning is returned.", "Existing access remains unchanged."], "Critical", "High Priority= 1-2 days")
add("2FA - Begin Setup", "Verify an authenticated user can begin authenticator setup.", "The user is signed in and 2FA is not yet enabled.", ["Open Settings.", "Start two-factor authentication setup.", "Inspect the JSON response."], ["The endpoint returns a QR code, secret, issuer, and recovery codes.", "Pending setup data is stored in the session.", "2FA is not finalized until verification succeeds."], "Critical", "High Priority= 1-2 days")
add("2FA - Setup Verification Validation", "Verify setup verification rejects an invalid authenticator code.", "Pending 2FA setup exists for the signed-in user.", ["Submit setup verification with an invalid authenticator code.", "Inspect the JSON response."], ["Verification fails.", "The response states the authenticator code is invalid.", "2FA remains disabled on the account."], "Critical", "High Priority= 1-2 days")
add("2FA - Setup Verification", "Verify setup verification enables 2FA with a valid authenticator code.", "Pending 2FA setup exists for the signed-in user.", ["Submit setup verification with a valid authenticator code.", "Inspect the response and stored account data."], ["Two-factor authentication is enabled.", "The secret, confirmation timestamp, and recovery codes are stored.", "The endpoint returns success."], "Critical", "High Priority= 1-2 days")
add("2FA - Disable", "Verify an authenticated user can disable their own 2FA setup.", "2FA is enabled on the signed-in account.", ["Submit the disable 2FA request.", "Inspect the account record afterward."], ["2FA status is disabled.", "Stored secret and recovery data are cleared.", "The endpoint returns success."], "Critical", "High Priority= 1-2 days")
add("Account - Delete Own Account", "Verify a non-admin user can soft-delete their own account with valid credentials.", "A signed-in non-admin account exists; if 2FA is enabled, a valid code is available.", ["Open the delete account flow.", "Enter the correct password.", "Enter a valid authenticator or recovery code when 2FA is enabled.", "Submit the deletion request."], ["The account is soft-deleted.", "Remember-me and 2FA session data are cleared.", "The user is signed out after deletion."], "Critical", "High Priority= 1-2 days")
add("Account - Delete Own Account Access Control", "Verify an Administrator cannot self-delete through the delete account endpoint.", "An Administrator account is signed in.", ["Open the delete account flow.", "Enter the correct password.", "Submit the request."], ["The request is denied.", "The response states an Administrator account cannot be deleted.", "The Administrator account remains active."], "Critical", "High Priority= 1-2 days")
add("Notifications - View", "Verify authenticated users can open the notification history page.", "The user is signed in.", ["Open Notifications.", "Review the transaction notifications list."], ["The notification history page loads.", "Notifications for the current user are shown.", "The empty state is handled without error."], "Moderate")
add("Notifications - Mark Read", "Verify a signed-in user can mark selected notifications as read.", "The user has unread app notifications.", ["Send a POST request to notifications/mark_read.php with valid IDs.", "Refresh the unread badge or feed."], ["The endpoint returns success.", "Selected notifications are marked read for the current user only.", "The unread count decreases."])
add("Contact - Non-Admin Compose", "Verify a non-admin can send a message through Contact or Inbox compose.", "A User, AA, or Editor account is signed in.", ["Open Contact or Inbox compose.", "Enter a subject and message.", "Leave recipients empty or choose admins.", "Submit the message."], ["The message is stored successfully.", "Admin recipients are derived by implemented non-admin recipient logic.", "The thread becomes visible in Inbox."])
add("Contact - Admin Compose", "Verify an Administrator can compose a message to all, admins, users, or specific user IDs.", "An Administrator account is signed in.", ["Open Inbox compose.", "Select recipients such as all, admins, users, or one specific user.", "Enter subject and message.", "Submit the message."], ["The message is saved successfully.", "Recipients are synchronized in contact_message_recipients.", "The intended recipients receive the thread."])
add("Contact - Validation", "Verify Contact rejects missing subject or missing content.", "The user is signed in.", ["Open Contact or Inbox compose.", "Leave subject blank, or leave both message and attachments empty.", "Submit the form."], ["The request is rejected.", "A validation error explains the missing fields.", "No message record is created."])
add("Inbox - User Thread Visibility", "Verify non-admin accounts only see their own or addressed threads.", "A non-admin account has at least one valid thread and at least one unrelated thread exists.", ["Sign in as a non-admin account.", "Open Inbox.", "Review the thread list."], ["Only threads owned by or addressed to the account are listed.", "Unrelated threads are not visible.", "Unread counts are scoped to the current user."], "Critical", "High Priority= 1-2 days")
add("Inbox - Admin Thread Visibility", "Verify Administrators can view admin-addressed mailbox threads.", "At least one thread exists that targets admins.", ["Sign in as Administrator.", "Open Inbox.", "Review the thread list and unread count."], ["Admin-addressed conversations are visible.", "The unread count reflects admin-targeted mail.", "Mailbox access remains authenticated-only."], "Critical", "High Priority= 1-2 days")
add("Inbox - Send Reply", "Verify a user can reply to a conversation they are allowed to access.", "The signed-in account can access an existing thread.", ["Open an accessible conversation.", "Enter reply text or attach a valid file.", "Submit the reply."], ["The reply is saved in contact_replies.", "Message read state updates for sender and recipients.", "The thread shows the new reply."])
add("Inbox - Reply Access Control", "Verify a user cannot reply to a conversation outside their mailbox scope.", "A thread exists that is not owned by or addressed to the signed-in non-admin account.", ["Send a reply request against the unrelated thread ID.", "Inspect the response."], ["The reply is rejected.", "The response states the user is not allowed to reply to the conversation.", "No reply record is created."], "Critical", "High Priority= 1-2 days")
add("Inbox - Edit Reply", "Verify a sender can edit a reply they are allowed to manage.", "The signed-in user owns a reply in an accessible thread.", ["Open the reply edit action.", "Change the reply content.", "Submit the update."], ["The reply text is updated.", "The updated timestamp changes.", "Unauthorized users cannot overwrite someone else’s reply."])
add("Inbox - Delete Message or Reply", "Verify permitted users can delete mailbox content under the implemented rules.", "A signed-in user owns a message or reply; an Administrator account also exists for admin-wide deletion testing.", ["Delete a message or reply as its permitted owner.", "Repeat the delete action as Administrator on content the admin can remove.", "Refresh the thread state."], ["Allowed deletions succeed.", "Mailbox state updates correctly.", "Unauthorized deletions are blocked."])
add("Calendar - View", "Verify any authenticated role can open the Calendar page.", "The account is signed in.", ["Sign in as Administrator, AA, Editor, and User.", "Open Calendar."], ["Calendar loads for each authenticated role.", "Existing events are fetched for the selected range.", "The page remains protected from unauthenticated access."], "Moderate")
add("Calendar - Create Event", "Verify an authenticated user can create a calendar event.", "The user is signed in.", ["Open Calendar.", "Create an event with a title and start date.", "Save the event."], ["The event is inserted with created_by set to the signed-in user.", "The calendar refreshes and shows the new event.", "A successful response is returned."])
add("Calendar - Update Access Control", "Verify a non-admin cannot update another user’s event.", "Two non-admin accounts exist and one of them has created an event.", ["Sign in as a different non-admin account.", "Send an update request for the other user’s event."], ["The update is denied with a 403 response.", "The event remains unchanged.", "Ownership rules are enforced by the endpoint."], "Critical", "High Priority= 1-2 days")
add("Calendar - Admin Update Override", "Verify an Administrator can update any existing event.", "An event exists that was created by another user.", ["Sign in as Administrator.", "Update the existing event."], ["The update succeeds.", "Event fields are changed as submitted.", "The admin override path works for existing events."], "Major", "High Priority= 1-2 days")
add("Calendar - Delete Access Control", "Verify non-admin deletion is limited to own events while admins can delete any event.", "One event exists for the current user and one event exists for another user.", ["As a non-admin, delete the current user’s own event.", "As the same non-admin, try to delete another user’s event.", "As Administrator, delete the remaining event."], ["The current user’s own event can be deleted.", "The non-admin cannot delete another user’s event.", "The Administrator can delete any event."], "Critical", "High Priority= 1-2 days")
add("Access Control - User vs Operations", "Verify the User role is blocked from operations pages guarded by auth_can_view_operations().", "A signed-in User account exists.", ["Sign in as User.", "Open pages/data-tracking-in directly.", "Repeat with pages/payout and pages/fund-monitoring."], ["Access is denied for operations pages.", "The user is redirected or shown a protected-page warning.", "The User role does not gain operations workspace access."], "Critical", "High Priority= 1-2 days")
add("Tracking - Incoming View Access", "Verify Administrator, AA, and Editor can open the Incoming tracking page.", "Accounts exist for Administrator, AA, and Editor.", ["Sign in as each non-user operations role.", "Open pages/data-tracking-in."], ["The Incoming page loads for Administrator, AA, and Editor.", "The page stays unavailable to the User role."], "Major", "High Priority= 1-2 days")
add("Tracking - Incoming Create", "Verify an allowed operations role can track an incoming document.", "Administrator, AA, or Editor is signed in.", ["Submit date_received and description.", "Optionally attach a valid file type.", "Save the record."], ["A new incoming record is created.", "A generated tracking number is assigned.", "The response confirms successful tracking."], "Major", "High Priority= 1-2 days")
add("Tracking - Incoming Validation", "Verify incoming tracking rejects missing required fields or unsupported file types.", "Administrator, AA, or Editor is signed in.", ["Submit the incoming form without date_received or description.", "Repeat with an unsupported file type."], ["Validation errors are returned.", "The record is not created when required data is missing or file type is invalid."])
add("Tracking - Incoming Update", "Verify an allowed operations role can update an existing incoming document.", "An incoming record exists and the user is Administrator, AA, or Editor.", ["Open the edit action for the incoming record.", "Change document fields or file handling options.", "Submit the update."], ["The incoming record is updated.", "File replacement or removal follows the implemented logic.", "A success message is returned."], "Major", "High Priority= 1-2 days")
add("Tracking - Forward Incoming Document", "Verify an allowed operations role can forward an incoming document to outgoing.", "An incoming record exists and has not yet been forwarded.", ["Open the forward action.", "Enter receiving office and forwarding date.", "Submit the forward request."], ["A matching outgoing record is created.", "The incoming record status becomes Forwarded.", "Duplicate forwarding is blocked afterward."], "Major", "High Priority= 1-2 days")
add("Tracking - Outgoing View Access", "Verify Administrator, AA, and Editor can open the Outgoing tracking page.", "Accounts exist for Administrator, AA, and Editor.", ["Sign in as each non-user operations role.", "Open pages/data-tracking-out."], ["The Outgoing page loads for Administrator, AA, and Editor.", "The page stays unavailable to the User role."], "Major", "High Priority= 1-2 days")
add("Tracking - Outgoing Create", "Verify an allowed operations role can create an outgoing tracking record.", "Administrator, AA, or Editor is signed in.", ["Enter date_out, description, receiving_office, and supported fields.", "Submit with a valid attachment or without a file when matching incoming data exists."], ["A new outgoing record is created.", "The record receives a generated tracking number.", "If no new upload is provided, a matching incoming file can be reused when available."], "Major", "High Priority= 1-2 days")
add("Tracking - Outgoing Validation", "Verify outgoing tracking rejects missing required fields or unsupported file types.", "Administrator, AA, or Editor is signed in.", ["Submit the outgoing form without date_out or description.", "Repeat with an unsupported file type."], ["Validation errors are returned.", "The record is not created when required data is missing or file type is invalid."])
add("Tracking - Outgoing Update", "Verify an allowed operations role can update an existing outgoing record.", "An outgoing record exists and the user is Administrator, AA, or Editor.", ["Open the outgoing edit action.", "Change document data or file handling flags.", "Submit the update."], ["The outgoing record is updated successfully.", "File replacement or removal follows the implemented rules.", "A success response is returned."], "Major", "High Priority= 1-2 days")
add("Partner-Beneficiaries - MEB View", "Verify non-user roles can open the MEB page under Partner-Beneficiaries.", "Administrator, AA, and Editor accounts exist.", ["Sign in as Administrator, AA, and Editor.", "Open pages/data-tracking-meb."], ["The MEB page loads for non-user roles.", "The User role cannot open the operations page."], "Major", "High Priority= 1-2 days")
add("Partner-Beneficiaries - MEB Export", "Verify the Master List export works for allowed operations roles.", "A non-user role is signed in and selected_year is set.", ["Open export_meb.php.", "Download the generated workbook."], ["The workbook is generated for the selected fiscal year.", "Exported rows are ordered as implemented.", "The export remains protected behind authentication."], "Moderate")
add("Partner-Beneficiaries - Bulk Action Admin Delete", "Verify an Administrator can run the MEB bulk delete action.", "Administrator is signed in and multiple MEB rows exist.", ["Select multiple MEB rows.", "Trigger bulk action with delete."], ["Selected rows are deleted.", "The page reports success.", "Bulk delete notifications and broadcasts are triggered."], "Major", "High Priority= 1-2 days")
add("Partner-Beneficiaries - Bulk Action Access Control", "Verify non-admin roles cannot execute the MEB bulk delete action.", "AA or Editor is signed in and MEB rows exist.", ["Submit bulk_action.php with action=delete and selected row IDs.", "Inspect the response."], ["Access is denied for non-admin roles.", "The selected rows remain unchanged."], "Critical", "High Priority= 1-2 days")
add("MEB Validation - Admin Manage", "Verify an Administrator can update MEB validation status.", "Administrator is signed in and target MEB rows exist.", ["Open the validation workflow or call update_validation_status.php.", "Submit one or more MEB row IDs with validated or not_validated status."], ["Validation values are updated for the selected rows.", "Audit and notification entries are produced.", "The endpoint returns success."], "Major", "High Priority= 1-2 days")
add("MEB Validation - Access Control", "Verify non-admin roles cannot access or export the MEB validation workflow.", "AA or Editor is signed in.", ["Open pages/data-tracking-meb-validation.", "Open export_meb_validation.php.", "Submit update_validation_status.php."], ["The validation workflow is blocked for non-admin roles.", "Validation export is denied.", "Validation update requests return access denied."], "Critical", "High Priority= 1-2 days")
add("Payout - View Access", "Verify Administrator and AA can open the payout page.", "Accounts exist for Administrator and AA.", ["Sign in as Administrator and AA.", "Open pages/payout."], ["The payout page loads for Administrator and AA.", "The User role is blocked by operations access.", "Access matches the implemented page and update behavior."], "Major", "High Priority= 1-2 days")
add("Payout - Update Record", "Verify Administrator and AA can update payout details.", "A payout breakdown record exists and selected_year has valid project variables.", ["Sign in as Administrator or AA.", "Edit a payout record or municipality group.", "Submit the update."], ["The payout update succeeds for Administrator and AA.", "Amounts are recalculated using configured project variables.", "Audit details are written for the update."], "Major", "High Priority= 1-2 days")
add("Payout - Update Access Control", "Verify Editor and User cannot update payout records.", "Editor and User accounts exist and a payout record exists.", ["Sign in as Editor.", "Submit update_payout.php or update_payout_group.php.", "Repeat as User."], ["Both roles are denied by the payout update handlers.", "Existing payout data remains unchanged."], "Critical", "High Priority= 1-2 days")
add("Payout - Export", "Verify the payout export generates a workbook when project variables exist.", "selected_year is set and daily_wage_rate is configured.", ["Open payout_export.php as an allowed operations role.", "Download the generated workbook."], ["The export is generated for the selected fiscal year.", "Totals and amount fields are computed correctly.", "Missing project variables block export with an error."], "Moderate")
add("Fund Monitoring - View", "Verify non-user roles can open the Fund Monitoring page.", "Administrator, AA, and Editor accounts exist.", ["Sign in as Administrator, AA, and Editor.", "Open pages/fund-monitoring."], ["The page loads for non-user roles through operations access.", "The User role is blocked from the operations workspace."], "Major")
add("Fund Monitoring - Admin Manage", "Verify only Administrator can save fund monitoring object codes, items, and month entries.", "Administrator is signed in and selected_year is set.", ["Submit save_fund_monitoring.php with action=save_object_code.", "Submit save_fund_monitoring.php with action=save_item.", "Submit save_fund_monitoring.php with action=save_month_entries."], ["Each save action succeeds for Administrator.", "Flash messages confirm the saved data.", "Records persist for the selected fiscal year."], "Major", "High Priority= 1-2 days")
add("Fund Monitoring - Access Control", "Verify AA and Editor can view Fund Monitoring but cannot modify the data.", "AA and Editor accounts are signed in separately.", ["Open pages/fund-monitoring as AA and Editor.", "Submit save_fund_monitoring.php using any supported action."], ["The page loads for AA and Editor.", "Save actions are rejected with the admin-only message.", "Existing fund monitoring data remains unchanged."], "Critical", "High Priority= 1-2 days")
add("Baseline Targets - View vs Manage", "Verify Baseline Targets management is limited to Administrator and Implementation Editor.", "Accounts exist for Administrator, Editor, AA, and User.", ["Sign in as Administrator and Editor and open implementation-status/program-targets.", "Repeat with AA and User."], ["Administrator and Editor can access the Baseline Targets page with management controls.", "AA does not gain Baseline Targets management access.", "User cannot manage Baseline Targets."], "Critical", "High Priority= 1-2 days")
add("Baseline Targets - Import", "Verify Administrator and Editor can import valid Baseline Targets Excel files.", "Administrator or Editor is signed in and has a valid target workbook.", ["Open Baseline Targets.", "Upload a valid .xls or .xlsx file through import-project-targets.php."], ["The import completes successfully.", "Target rows are inserted or updated for the selected fiscal year.", "No access error is returned for Administrator or Editor."], "Major", "High Priority= 1-2 days")
add("Baseline Targets - Import Validation", "Verify target import rejects unsupported files or invalid workbook structure.", "Administrator or Editor is signed in.", ["Upload a non-Excel file.", "Upload an Excel file with missing required headers or invalid row values."], ["Invalid file types are rejected.", "Invalid workbook structures trigger import error feedback.", "No malformed data is imported."])
add("Baseline Targets - Save", "Verify Administrator and Editor can save a valid Baseline Target record.", "Administrator or Editor is signed in and selected_year is set.", ["Submit save-project-target.php with valid location data and entries.", "Review the JSON response and stored record."], ["The target record is saved successfully.", "LAWA and BINHI counts are computed from the submitted entries.", "The selected fiscal year is used for storage."], "Major", "High Priority= 1-2 days")
add("Baseline Targets - Delete", "Verify Administrator and Editor can delete a Baseline Target record.", "A Baseline Target exists for the selected fiscal year.", ["Sign in as Administrator or Editor.", "Submit delete-project-target.php with the target ID."], ["The target row is deleted when it matches the selected fiscal year.", "A success JSON response is returned.", "Invalid or missing target IDs are rejected."], "Major", "High Priority= 1-2 days")
add("Baseline Targets - Access Control", "Verify AA and User cannot save, import, or delete Baseline Targets.", "AA and User accounts exist.", ["Sign in as AA and submit save-project-target.php, import-project-targets.php, and delete-project-target.php.", "Repeat as User."], ["Access is denied for AA and User on target management endpoints.", "No Baseline Target data is changed."], "Critical", "High Priority= 1-2 days")
add("Program Activities - View", "Verify all authenticated roles can open Program Activities, with AA and User limited to view mode.", "Accounts exist for Administrator, AA, Editor, and User.", ["Sign in as each role.", "Open implementation-status/program-activities."], ["The Program Activities page loads for all authenticated roles.", "Administrator and Editor receive management capabilities.", "AA and User remain in view-only mode."], "Critical", "High Priority= 1-2 days")
add("Program Activities - Save", "Verify Administrator and Editor can save Program Activities data.", "Administrator or Editor is signed in and selected_year is set.", ["Submit save-imp-status.php with valid municipality metadata and rows.", "Inspect the JSON response and stored records."], ["Program activity metadata and rows are saved.", "The upsert path updates existing fiscal-year metadata when needed.", "A success response is returned."], "Major", "High Priority= 1-2 days")
add("Program Activities - Access Control", "Verify AA and User cannot save Program Activities changes.", "AA and User accounts exist.", ["Sign in as AA and submit save-imp-status.php with valid payload data.", "Repeat as User."], ["Access is denied for AA and User.", "Existing Program Activities data remains unchanged."], "Critical", "High Priority= 1-2 days")
add("Project Location Maps - View", "Verify all authenticated roles can open Project Location Maps.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as each role.", "Open implementation-status/project-location-maps."], ["The Project Location Maps page loads for all authenticated roles.", "Map data is shown for the selected filters and fiscal year."], "Moderate")
add("Project Location Records - View", "Verify all authenticated roles can open Project Location Records.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as each role.", "Open implementation-status/project-location-records."], ["The Project Location Records page loads for all authenticated roles.", "The page lists records generated from Program Activities data."], "Moderate")
add("LAWA Summary - View", "Verify all authenticated roles can open LAWA Summary.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as each role.", "Open implementation-status/lawa-summary."], ["The LAWA Summary page loads for all authenticated roles.", "Summary data is filtered by the selected fiscal year."], "Moderate")
add("BINHI Summary - View", "Verify all authenticated roles can open BINHI Summary.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as each role.", "Open implementation-status/binhi-summary."], ["The BINHI Summary page loads for all authenticated roles.", "Summary data is filtered by the selected fiscal year."], "Moderate")
add("Reports - Sectoral Summary and Export", "Verify all authenticated roles can open Sectoral Summary and export it.", "selected_year is set and the account is signed in.", ["Open pages/summary/sectoral.", "Trigger the sectoral export endpoint."], ["Sectoral Summary loads for the signed-in user.", "The export workbook is generated for the selected fiscal year."], "Moderate")
add("Reports - Partner-Beneficiaries Profile Access", "Verify Partner-Beneficiaries Profile is available only to non-user roles.", "Accounts exist for Administrator, AA, Editor, and User.", ["Sign in as Administrator, AA, and Editor and open pages/summary/beneficiary-profile.", "Repeat as User.", "Open the profile export endpoint as a non-user role."], ["Administrator, AA, and Editor can view the profile page.", "The User role is denied access to the profile page.", "The export generates for allowed non-user roles."], "Critical", "High Priority= 1-2 days")
add("Reports - PWD Summaries", "Verify all authenticated roles can open the PWD Disability Summary and Sex Disaggregation reports.", "Accounts exist for Administrator, AA, Editor, and User.", ["Sign in as each role.", "Open pages/summary/pwd/pwd.", "Open pages/summary/pwd/sex-disaggregated-pwd."], ["Both PWD report pages load for authenticated roles.", "Report data respects the selected fiscal year."], "Moderate")
add("Utilities - Crossmatching View", "Verify authenticated users can open the Crossmatching tool.", "An authenticated account exists.", ["Sign in as Administrator, AA, Editor, and User.", "Open the crossmatch directory."], ["The Crossmatching tool is reachable for authenticated users.", "Unauthenticated access stays blocked by the shared auth flow."], "Moderate")
add("Utilities - Deduplication View", "Verify authenticated users can open the Deduplication tool.", "An authenticated account exists.", ["Sign in as Administrator, AA, Editor, and User.", "Open the deduplication directory."], ["The Deduplication tool is reachable for authenticated users.", "Unauthenticated access stays blocked by the shared auth flow."], "Moderate")
add("Utilities - MEBIS Name-Matching Template Access", "Verify the MEBIS Name-Matching Template generator is Administrator-only.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as Administrator and open mebis-consolidator/.", "Repeat as AA, Editor, and User."], ["Administrator can open the generator.", "AA, Editor, and User are denied by admin-generator access enforcement."], "Critical", "High Priority= 1-2 days")
add("Utilities - MEB Import Template Access", "Verify the MEB Import Template generator is Administrator-only.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as Administrator and open mebis-lgu-template/.", "Repeat as AA, Editor, and User."], ["Administrator can open the generator.", "AA, Editor, and User are denied by admin-generator access enforcement."], "Critical", "High Priority= 1-2 days")
add("Administration - Users Management", "Verify an Administrator can open Users Management and change a user role.", "Administrator is signed in and a non-admin account exists.", ["Open admin/users_management.", "Change the target user type through change_user_type.php."], ["Users Management loads for Administrator.", "The target user type is updated successfully.", "A non-admin attempting the same action would be denied."], "Critical", "High Priority= 1-2 days")
add("Administration - Users Management Access Control", "Verify non-admin roles cannot access Users Management.", "AA, Editor, and User accounts exist.", ["Sign in as AA, Editor, and User.", "Open admin/users_management.", "Attempt to call admin/change_user_type.php."], ["The page is blocked for non-admin roles.", "The change role action returns access denied."], "Critical", "High Priority= 1-2 days")
add("Administration - Deactivate and Restore Users", "Verify an Administrator can deactivate and restore user accounts.", "Administrator is signed in and a non-admin account exists.", ["Deactivate a user through the admin deactivate action.", "Open the restore workflow.", "Restore the deactivated user."], ["The target account is soft-deactivated.", "The restore workflow returns the account to active use.", "The actions remain Administrator-only."], "Critical", "High Priority= 1-2 days")
add("Administration - Password Security", "Verify only Administrator can run password reminder and forced reset actions.", "Administrator, AA, Editor, and User accounts exist and a target user exists.", ["Sign in as Administrator and trigger send_reminder or force_reset through password_security_action.php.", "Repeat as AA or Editor."], ["Administrator actions succeed and log the security event.", "AA and Editor are denied access."], "Critical", "High Priority= 1-2 days")
add("Administration - Project Variables Manage", "Verify Administrator and Editor can manage Project Variables.", "Administrator and Editor accounts exist and selected_year is set.", ["Open admin/project_variables as Administrator and Editor.", "Save or update a valid project variable through save_project_variable.php."], ["The page loads for Administrator and Editor.", "Valid project variable changes are saved successfully.", "Duplicate fiscal year and key combinations are prevented by the save result."], "Major", "High Priority= 1-2 days")
add("Administration - Project Variables Access Control", "Verify AA and User cannot manage Project Variables.", "AA and User accounts exist.", ["Sign in as AA or User.", "Open admin/project_variables.", "Submit save_project_variable.php."], ["AA and User are denied Project Variables management access.", "No project variable data is changed."], "Critical", "High Priority= 1-2 days")
add("Administration - Maintenance Mode", "Verify only Administrator can open and save Maintenance Mode settings.", "Administrator, AA, Editor, and User accounts exist.", ["Sign in as Administrator and open admin/maintenance.", "Save maintenance settings through admin/save_maintenance_settings.php.", "Repeat page access as AA, Editor, and User."], ["Administrator can open and save maintenance settings.", "AA, Editor, and User are denied access to the maintenance tool.", "Maintenance changes use the protected admin save path only."], "Critical", "High Priority= 1-2 days")


def build_workbook():
    workbook = load_workbook(TEMPLATE)
    main = workbook["Dromic 1.0 Test case format"]

    style_cache = {}
    for col in range(1, 24):
        cell = main.cell(7, col)
        style_cache[col] = {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        }

    for sheet_name in list(workbook.sheetnames):
        if sheet_name != "Dromic 1.0 Test case format":
            workbook.remove(workbook[sheet_name])

    main.title = "KODUS Test Cases"
    main.sheet_state = "visible"
    workbook.active = 0

    main["B1"] = "KODUS WEB APP"
    main["B2"] = "PHP Web Application / Role-based access / Fiscal year data"
    main["B3"] = "Functional, Validation, and Access Control Coverage"
    main["B4"] = "Authentication, Dashboard, Settings, 2FA, Messaging, Calendar, Tracking, Implementation Status, Reports, Utilities, Administration"

    if main.max_row >= 7:
        main.delete_rows(7, main.max_row - 6)

    for index, case in enumerate(CASES, start=1):
        row = 6 + index
        values = {
            1: index,
            2: f"KODUS-TC-{index:03d}",
            3: case["module"],
            4: case["procedure"],
            5: case["expected"],
            6: "",
            7: case["severity"],
            8: case["priority"],
            9: "",
            10: "",
            11: "",
            12: "",
            13: "",
            14: "",
        }

        for col in range(1, 24):
            cell = main.cell(row, col)
            style = style_cache[col]
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy(style["protection"])
            if col in values:
                cell.value = values[col]

        for col in [3, 4, 5, 6, 10, 11]:
            main.cell(row, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        main.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
        main.cell(row, 2).alignment = Alignment(horizontal="center", vertical="top")

        max_lines = max(case["procedure"].count("\n") + 1, case["expected"].count("\n") + 1)
        main.row_dimensions[row].height = min(max(42, max_lines * 13), 150)

    last_row = 6 + len(CASES)
    main.print_title_rows = "1:6"
    main.print_area = f"A1:N{last_row}"
    main.page_setup.orientation = "landscape"
    main.page_setup.fitToWidth = 1
    main.page_setup.fitToHeight = 0
    main.sheet_properties.pageSetUpPr.fitToPage = True
    main.sheet_view.zoomScale = 85

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    output = build_workbook()
    print(output)
    print(f"cases={len(CASES)}")
