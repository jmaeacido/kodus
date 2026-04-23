from copy import copy
from pathlib import Path
import math

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parent
TEMPLATE_PATH = Path(r"C:\Users\jmaeacido\Downloads\FAITH User Story.xlsx")
OUTPUT_XLSX = ROOT / "KODUS_User_Story.xlsx"


SECTIONS = [
    {
        "role": "Administrator",
        "summary": (
            "Primary responsibilities: full system control, user and account governance, "
            "security policy administration, audit oversight, app settings, and unrestricted "
            "access across operational and implementation workspaces."
        ),
        "stories": [
            (
                "Admin Login and Full Workspace Access",
                "As an Administrator, I want to sign in with elevated access so that I can supervise every protected KODUS workspace and administration page.",
                "Covers authenticated access plus shared security flows such as password updates and 2FA setup in personal settings.",
            ),
            (
                "Manage Users",
                "As an Administrator, I want to create, reclassify, activate, and deactivate user accounts so that each person receives the correct role and access level.",
                "Admin-only user management pages and role-change actions are enforced through page guards and admin transactions.",
            ),
            (
                "Restore Deleted Users",
                "As an Administrator, I want to restore deleted user accounts so that valid access can be recovered without rebuilding records from scratch.",
                "Admin-only restore pages are available in the administration workspace.",
            ),
            (
                "Manage Security Controls",
                "As an Administrator, I want to manage password security settings and privileged recovery actions so that the platform stays secure and supportable.",
                "Aligned with admin-only password security tools and privileged recovery actions such as user 2FA resets.",
            ),
            (
                "Manage App Settings and Maintenance",
                "As an Administrator, I want to update maintenance mode and application-level settings so that KODUS remains secure, available, and properly configured.",
                "Backed by admin-only maintenance controls and administrator settings pages.",
            ),
            (
                "Review Audit Activity",
                "As an Administrator, I want to review audit-related activity for sensitive changes so that I can trace actions and investigate issues when needed.",
                "Privileged updates write audit entries; audit review remains an administrator oversight function.",
            ),
            (
                "Manage Project Variables",
                "As an Administrator, I want to maintain project variables by fiscal year so that computed implementation summaries and payout values stay correct.",
                "Allowed through `auth_can_manage_project_variables()` and the admin project variables workspace.",
            ),
            (
                "Manage Baseline Targets",
                "As an Administrator, I want to add, import, edit, and delete baseline targets so that approved planning values stay accurate.",
                "Target create, import, edit, and delete transactions require `auth_can_manage_program_targets()`.",
            ),
            (
                "Manage Implementation Status and Program Activities",
                "As an Administrator, I want to edit implementation activity records and accomplishment details so that official status reporting stays current.",
                "Implementation save actions require `auth_can_manage_program_activities()`; admin inherits the full editor capability set.",
            ),
            (
                "Manage Tracking and Payout Transactions",
                "As an Administrator, I want to log, edit, forward, and update tracking and payout records so that operational document movement and payout monitoring remain controlled end to end.",
                "Incoming, outgoing, and payout transaction buttons are explicitly available to `admin`; report pages also provide export actions.",
            ),
        ],
    },
    {
        "role": "Administrative Staff (AA)",
        "summary": (
            "Primary responsibilities: operational support, inbox and notification handling, "
            "view-only use of implementation pages, and management of only the tracking and payout "
            "transactions explicitly allowed by code."
        ),
        "stories": [
            (
                "View Operations Dashboard",
                "As an Administrative Staff member, I want to open the dashboard and operations workspace so that I can monitor daily document, message, and support activity.",
                "AA is included in non-user operations access through `auth_can_view_operations()`.",
            ),
            (
                "Manage Notifications",
                "As an Administrative Staff member, I want to review and clear notifications so that my operational queue stays organized.",
                "Notifications are available to authenticated users through the notifications center and mark-read endpoint.",
            ),
            (
                "Manage Inbox and Contact Replies",
                "As an Administrative Staff member, I want to send contact messages and reply inside the inbox so that concerns can be acknowledged and coordinated quickly.",
                "Inbox and contact flows are available to authenticated users; non-admin users operate within their accessible threads.",
            ),
            (
                "View Implementation Status",
                "As an Administrative Staff member, I want to review implementation status pages without editing them so that I can support operations using current program information.",
                "Program activities page exposes viewer mode when the role does not pass the implementation-edit permission checks.",
            ),
            (
                "View Baseline Targets",
                "As an Administrative Staff member, I want to review baseline target data without changing it so that I can reference official targets during coordination work.",
                "Target add, import, edit, and delete actions remain restricted to administrators and implementation editors.",
            ),
            (
                "View Program Summaries",
                "As an Administrative Staff member, I want to view LAWA and BINHI summaries so that I can answer operational questions using current implementation outputs.",
                "Summary pages are available when logged in and do not expose AA editing transactions.",
            ),
            (
                "View Project Location Pages",
                "As an Administrative Staff member, I want to open project location records and maps so that I can verify where projects are being implemented.",
                "Project location pages provide read access; no AA edit controls are exposed there.",
            ),
            (
                "View Tracking Records",
                "As an Administrative Staff member, I want to review incoming, outgoing, and payout tracking pages before acting so that I can confirm the current document or payout status.",
                "Tracking pages are part of the operations menu for non-user roles.",
            ),
            (
                "Manage Incoming Tracking Records",
                "As an Administrative Staff member, I want to log and edit incoming document records so that received documents remain complete, searchable, and ready for routing.",
                "Track and edit controls in `pages/data-tracking-in.php` are explicitly shown only for `admin` and `aa`.",
            ),
            (
                "Forward Incoming Documents",
                "As an Administrative Staff member, I want to forward incoming documents to the next office or personnel so that the routing trail is properly recorded.",
                "Forward actions in incoming tracking are limited to `admin` and `aa` when the document is not yet forwarded.",
            ),
            (
                "Manage Outgoing Tracking Records",
                "As an Administrative Staff member, I want to log and edit outgoing document records so that released documents carry complete routing details.",
                "Track and edit controls in `pages/data-tracking-out.php` are explicitly limited to `admin` and `aa`.",
            ),
            (
                "Manage Payout Records",
                "As an Administrative Staff member, I want to update only the payout records I am allowed to handle so that payout monitoring remains accurate.",
                "Both `pages/update_payout.php` and `pages/update_payout_group.php` authorize only `admin` and `aa` for payout updates.",
            ),
        ],
    },
    {
        "role": "Implementation Editor",
        "summary": (
            "Primary responsibilities: maintaining implementation data for fiscal-year variables, "
            "baseline targets, and activity records, while using view access for summaries, reports, "
            "and project location outputs."
        ),
        "stories": [
            (
                "Access the Editor Workspace",
                "As an Implementation Editor, I want the dashboard to surface editor tools and quick actions so that I can go directly to the data-maintenance pages I own.",
                "The home dashboard renders editor-specific workspace content and links when the current role is `editor`.",
            ),
            (
                "Manage Project Variables",
                "As an Implementation Editor, I want to maintain fiscal-year project variables so that computed program summaries and related implementation values stay correct.",
                "Editors share `auth_can_manage_project_variables()` access with administrators.",
            ),
            (
                "Manage Baseline Targets",
                "As an Implementation Editor, I want to add, import, edit, and delete baseline targets so that approved target data is complete before execution.",
                "Target maintenance transactions are enabled only for administrators and implementation editors.",
            ),
            (
                "Manage Implementation Status",
                "As an Implementation Editor, I want to update implementation status records so that accomplishment figures match actual field progress.",
                "Implementation save actions are protected by `auth_can_manage_program_activities()`.",
            ),
            (
                "Manage Program Activities",
                "As an Implementation Editor, I want to maintain schedules, coverage rows, and accomplishment details so that official activity records remain accurate.",
                "Editors receive the activity edit path while non-managers only see viewer mode.",
            ),
            (
                "View LAWA Summary",
                "As an Implementation Editor, I want to review the LAWA summary after updates so that I can confirm the data appears correctly in outputs.",
                "LAWA summary is viewable when logged in and reflects implementation data maintained upstream.",
            ),
            (
                "View BINHI Summary",
                "As an Implementation Editor, I want to review the BINHI summary after updates so that I can validate the resulting totals and computed indicators.",
                "BINHI summary is viewable when logged in and depends on the maintained implementation dataset.",
            ),
            (
                "View Project Location Records and Map",
                "As an Implementation Editor, I want to inspect project location record and map pages so that I can verify the location outputs generated from implementation data.",
                "Project location pages provide read access for logged-in users; no separate editor-only map transaction is implemented.",
            ),
            (
                "Generate and Export Reports",
                "As an Implementation Editor, I want to export summaries and implementation reports so that program updates can be shared outside the system.",
                "Reporting pages expose export actions such as Excel export on summary views.",
            ),
        ],
    },
    {
        "role": "User",
        "summary": (
            "Primary responsibilities: limited self-service access to dashboard, personal settings, "
            "notifications, messages, summaries, and project-location views without access to the "
            "operations workspace or implementation editing transactions."
        ),
        "stories": [
            (
                "Secure User Login",
                "As a User, I want to sign in with my account so that I can access the limited KODUS workspace assigned to me.",
                "Users can authenticate normally but are excluded from the operations workspace by `auth_can_view_operations()`.",
            ),
            (
                "View Dashboard",
                "As a User, I want to open the dashboard so that I can see my available summaries, messages, and limited navigation options.",
                "The user dashboard stays separate from admin and editor workspace views.",
            ),
            (
                "Manage Personal Profile and 2FA",
                "As a User, I want to update my own settings, password, and 2FA preferences so that I can keep my account secure without needing administrator help.",
                "Backed by `settings.php`, `begin_2fa_setup.php`, `verify_2fa_code.php`, and `disable_2fa.php`.",
            ),
            (
                "View Notifications",
                "As a User, I want to review my notifications and clear items I have already handled so that my activity feed stays organized.",
                "Notification viewing and mark-read actions are available to authenticated users.",
            ),
            (
                "Start Inbox Conversations",
                "As a User, I want to send messages through the contact and inbox features so that I can start conversations with administrators or other available users when I need help or coordination.",
                "The inbox compose list loads user accounts for non-admin senders, and non-admin contact submissions are stored as internal inbox threads.",
            ),
            (
                "View and Reply to Accessible Inbox Threads",
                "As a User, I want to read and reply to the conversation threads addressed to me or started by me so that I can continue discussions with other participants inside KODUS.",
                "Non-admin inbox access stays limited to the message threads visible to that user, but those threads can include other user participants.",
            ),
            (
                "View LAWA and BINHI Summaries",
                "As a User, I want to review high-level LAWA and BINHI summaries so that I can understand program results without changing any source data.",
                "Summary access is read-only from the user role perspective.",
            ),
            (
                "View Project Location Records and Map",
                "As a User, I want to inspect project location records and map pages so that I can see where implementations are taking place.",
                "Location pages are available for viewing but do not expose user edit controls.",
            ),
        ],
    },
]


def estimate_row_height(texts, base=24, per_line=14):
    lines = 1
    for text in texts:
        if not text:
            continue
        raw_lines = str(text).splitlines() or [""]
        for raw in raw_lines:
            width_hint = max(1, len(raw))
            lines += max(0, math.ceil(width_hint / 52) - 1)
    return max(base, min(90, base + (lines - 1) * per_line))


def apply_style(target, source):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def build_workbook():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["FAITH WEB"]

    for sheet_name in list(wb.sheetnames):
        if sheet_name != "FAITH WEB":
            del wb[sheet_name]

    ws.title = "KODUS WEB"

    style_a_header = copy(ws["A12"])
    style_role_header = copy(ws["B12"])
    style_section_note = copy(ws["D12"])
    style_num = copy(ws["B13"])
    style_feature = copy(ws["C13"])
    style_story = copy(ws["D13"])
    style_note = copy(ws["E13"])
    style_blank_a = copy(ws["A13"])

    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 12 or merged.max_row >= 12:
            ws.unmerge_cells(str(merged))

    if ws.max_row > 11:
        ws.delete_rows(12, ws.max_row - 11)

    ws["A2"] = "KODUS SYSTEM"
    ws["D4"] = "KliMalasakit Online Document Updating System"
    ws["D5"] = "KODUS Administration"
    ws["A10"] = (
        "KliMalasakit Online Document Updating System (KODUS) is a web-based platform for "
        "user administration, document tracking, implementation-status reporting, location views, "
        "notifications, and internal messaging. The user stories below are based on implemented "
        "role checks, page guards, sidebar visibility, and transaction permissions in the codebase."
    )
    ws["A11"] = (
        "Access Matrix Basis: validated from PHP role checks, page guards, sidebar/menu visibility, "
        "transaction endpoints, and tracking-page permissions as implemented in the current KODUS codebase."
    )

    current_row = 12
    for section in SECTIONS:
        role = section["role"]
        summary = section["summary"]
        stories = section["stories"]

        section_start = current_row
        section_end = current_row + len(stories)

        ws.merge_cells(start_row=section_start, start_column=1, end_row=section_end, end_column=1)
        ws.merge_cells(start_row=section_start, start_column=2, end_row=section_start, end_column=3)
        ws.merge_cells(start_row=section_start, start_column=4, end_row=section_start, end_column=5)

        apply_style(ws.cell(section_start, 1), style_a_header)
        apply_style(ws.cell(section_start, 2), style_role_header)
        apply_style(ws.cell(section_start, 4), style_section_note)

        ws.cell(section_start, 2).value = role
        ws.cell(section_start, 4).value = summary
        ws.row_dimensions[section_start].height = estimate_row_height([role, summary], base=36, per_line=12)

        for index, (feature, story, note) in enumerate(stories, start=1):
            row = section_start + index
            apply_style(ws.cell(row, 1), style_blank_a)
            apply_style(ws.cell(row, 2), style_num)
            apply_style(ws.cell(row, 3), style_feature)
            apply_style(ws.cell(row, 4), style_story)
            apply_style(ws.cell(row, 5), style_note)

            ws.cell(row, 2).value = float(index)
            ws.cell(row, 3).value = feature
            ws.cell(row, 4).value = story
            ws.cell(row, 5).value = note
            ws.row_dimensions[row].height = estimate_row_height([feature, story, note])

        current_row = section_end + 1

    ws.print_area = f"A1:E{current_row - 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    build_workbook()
